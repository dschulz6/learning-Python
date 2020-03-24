#!/Python37-32
#
# This program opens the "trip_record" spreadsheet from Farmers
# insurance and separates the business trips from the personal trips
# and calculates the mileage for each
#
from __future__ import print_function, unicode_literals
import openpyxl 

# Initialize mileage counter
biz_total = 0
my_total = 0
row_count = 0
#
# open spreadsheet called "trip_record.xlsx"
wb = openpyxl.load_workbook('trip_record.xlsx')
#
# open sheet in spreadsheet named "trip_record"
sheet = wb['trip_record']
#
# determine the number of rows in the spreadsheet
row_count = sheet.max_row
#
# print date range
start_date = sheet.cell(row=2, column=2).value
stop_date = sheet.cell(row=(row_count), column=3).value
print()
print()
print ("Start Date  ",(start_date))
print ()
print ("Stop Date  ",(stop_date))
# add values in 5th column (business or personal distance in miles)
for i in range(2, (row_count)+ 1, 1):
	if sheet.cell(row=i, column=8).value == 'Yes':
	    biz_total = sheet.cell(row=i, column=5).value + biz_total	    
	elif sheet.cell(row=i, column=8).value == 'No':
	    my_total = sheet.cell(row=i, column=5).value + my_total
#print total miles driven
print()	
print ("Business Miles = ",round(biz_total,2))
print ("Personal Miles = ",round(my_total,2))
print()


